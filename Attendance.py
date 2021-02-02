import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


def missed_list(filename, filename2):
    # Fills roster with section class roster
    roster = pd.read_excel("AttendanceSheet.xlsx")
    roster = roster.Names.tolist()

    # Fills attended with names of those who attended
    attended = pd.read_csv(filename)
    attended = attended.Name.tolist()
    attended2 = pd.read_csv(filename2)
    attended2 = attended2.Name.tolist()
    attended.extend(attended2)

    # Removes anyone who wasn't in class from roster
    for person in attended:
        if person in roster:
            roster.remove(person)

    return roster


missed = missed_list("Attendance.csv", "Attendance2.csv")
print(missed)

workbook = load_workbook('AttendanceSheet.xlsx')
sheet = workbook.active

class_roster = pd.read_excel("AttendanceSheet.xlsx")
class_roster = class_roster.Names.tolist()

def check_daily_attendance(c,missed):
    i = 2
    for person in class_roster:
        if person in missed:
            sheet.cell(row=i,column=c).fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
        else:
            sheet.cell(row=i, column=c).fill = PatternFill(fill_type='solid', start_color='32CD32', end_color='32CD32')
        i = i+1


check_daily_attendance(2, missed)


workbook.save("AttendanceSheet.xlsx")